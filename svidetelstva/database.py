import os
import sqlite3
import json
import io
import re
import unicodedata
from flask_login import UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

STUDENT_XLSX_HEADERS = [
    'Prezime', 'Ime', 'Ime na roditel', 'Paralelka', 'Roden/a na (god.)',
    'Mesto na raganje', 'Opstina na raganje', 'Drzava', 'Drzavjanstvo'
]


def _normalize_class_name(name):
    name = unicodedata.normalize('NFKC', str(name)).strip().upper()
    m = re.match(r'^([IVX]+)[\s_-]*(\d+)$', name)
    if m:
        return f'{m.group(1)}-{m.group(2)}'
    return name

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'svidetelstva.db')


class User(UserMixin):
    def __init__(self, id, username, role, full_name):
        self.id = str(id)
        self.username = username
        self.role = role
        self.full_name = full_name


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'teacher',
            full_name TEXT
        );

        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS user_classes (
            user_id INTEGER,
            class_id INTEGER,
            PRIMARY KEY (user_id, class_id),
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            parent_name TEXT,
            birth_date TEXT,
            birth_place TEXT,
            birth_municipality TEXT,
            birth_country TEXT DEFAULT 'Република Северна Македонија',
            citizenship TEXT DEFAULT 'македонско',
            class_id INTEGER,
            FOREIGN KEY (class_id) REFERENCES classes(id)
        );

        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            struka TEXT,
            profil TEXT
        );

        CREATE TABLE IF NOT EXISTS mk_cities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS behavior_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS countries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS certificates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            student_name TEXT,
            class_id INTEGER,
            school_year TEXT,
            attempt_number TEXT,
            completed_year_top TEXT,
            class_year TEXT,
            parent_name TEXT,
            birth_date TEXT,
            birth_place TEXT,
            birth_municipality TEXT,
            birth_country TEXT,
            citizenship TEXT,
            main_book_number TEXT,
            main_book_year TEXT,
            subjects_grades TEXT DEFAULT '[]',
            behavior TEXT,
            absences_justified TEXT,
            absences_unjustified TEXT,
            completed_year TEXT,
            education_type TEXT,
            area TEXT,
            profile TEXT,
            overall_grade TEXT,
            place TEXT,
            cert_date TEXT,
            del_br TEXT,
            verification_act TEXT,
            verification_date TEXT,
            verification_issued_by TEXT,
            status TEXT NOT NULL DEFAULT 'draft',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (class_id) REFERENCES classes(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS ministries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            act_number TEXT,
            act_date TEXT
        );
    ''')

    # Backward-compatible schema migration for older databases.
    cert_columns = {row['name'] for row in conn.execute("PRAGMA table_info(certificates)").fetchall()}
    missing_columns = [
        ('parent_name', 'TEXT'),
        ('birth_date', 'TEXT'),
        ('birth_place', 'TEXT'),
        ('birth_municipality', 'TEXT'),
        ('birth_country', 'TEXT'),
        ('citizenship', 'TEXT'),
        ('completed_year_top', 'TEXT'),
        ('main_book_year', 'TEXT'),
        ('class_teacher', 'TEXT'),
        ('director', 'TEXT')
    ]
    for col_name, col_type in missing_columns:
        if col_name not in cert_columns:
            conn.execute(f'ALTER TABLE certificates ADD COLUMN {col_name} {col_type}')

    if not conn.execute('SELECT id FROM users WHERE role="admin"').fetchone():
        conn.execute(
            'INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)',
            ('admin', generate_password_hash('admin123'), 'admin', 'Администратор')
        )

    # Ensure standard class set exists: I-1..I-15, II-1..II-15, III-1..III-15, IV-1..IV-15.
    for year in ['I', 'II', 'III', 'IV']:
        for idx in range(1, 16):
            conn.execute('INSERT OR IGNORE INTO classes (name) VALUES (?)', (f'{year}-{idx}',))

    first_behavior = conn.execute('SELECT name FROM behavior_types ORDER BY id LIMIT 1').fetchone()
    if not first_behavior or first_behavior['name'] != 'Примерно':
        conn.execute('DELETE FROM behavior_types')
        for b in ['Примерно', 'Добро', 'Незадоволително']:
            conn.execute('INSERT INTO behavior_types (name) VALUES (?)', (b,))

    if not conn.execute('SELECT id FROM subjects').fetchone():
        default_subjects = [
            'Македонски јазик и литература',
            'Математика',
            'Англиски јазик',
            'Германски јазик',
            'Француски јазик',
            'Италијански јазик',
            'Историја',
            'Географија',
            'Физика',
            'Хемија',
            'Биологија',
            'Информатика',
            'Спорт и спортски активности',
            'Ликовна уметност',
            'Музичка уметност',
            'Социологија',
            'Говорење и пишување',
            'Елементарна алгебра',
            'Елементарна алгебра и геометрија',
            'Етика',
            'Информатичка технологија',
            'Класичен јазик',
            'Алгебра',
            'Линеарна алгебра и аналитичка геометрија',
            'Програмски јазици',
            'Латински јазик',
            'Педагогија',
            'Вовед во правото',
            'Филозофија',
            'Бизнис и претприемништво',
            'Математична анализа',
            'Економија',
            'Менаџмент',
            'Логика',
            'Психологија',
            'Филозофија-2',
            'Компаративна книжевност',
            'Драмска уметност',
            'Руски јазик',
            'Албански јазик и литература',
            'Турски јазик и литература',
            'Српски Јазик и Литература',
            'Старогрчки јазик',
            'Шпански јазик',
            'Грчки јазик'
            'Образование за животни вештини'
        ]
        for s in default_subjects:
            conn.execute('INSERT INTO subjects (name) VALUES (?)', (s,))

    if not conn.execute('SELECT id FROM countries').fetchone():
        countries = [
            # Балкан и поранешна Југославија — прво
            'Република Северна Македонија', 'Република Србија', 'Република Хрватска',
            'Република Словенија', 'Босна и Херцеговина', 'Република Косово',
            'Република Албанија', 'Република Бугарија', 'Република Грција',
            'Република Црна Гора',
            # Европа
            'Федерална Република Германија', 'Република Австрија', 'Конфедерација Швајцарија',
            'Италијанска Република', 'Француска Република', 'Кралство Шпанија',
            'Португалска Република', 'Обединето Кралство Велика Британија и Северна Ирска',
            'Кралство Шведска', 'Кралство Норвешка', 'Кралство Данска',
            'Република Финска', 'Кралство Холандија', 'Кралство Белгија',
            'Велико Војводство Луксембург', 'Ирска', 'Кралство Шпанија',
            'Република Полска', 'Чешка Република', 'Словачка Република',
            'Унгарија', 'Романија', 'Република Молдавија', 'Република Белорусија',
            'Република Литванија', 'Република Летонија', 'Република Естонија',
            'Руска Федерација', 'Украина', 'Република Исланд', 'Република Малта',
            'Република Кипар', 'Кнежевство Монако', 'Република Сан Марино',
            'Република Андора', 'Република Португалија', 'Република Хрватска',
            'Република Косово', 'Република Македонија',
            # Азија
            'Република Турција', 'Сирија', 'Либан', 'Израел', 'Јорданија',
            'Саудиска Арабија', 'Обединети Арапски Емирати', 'Катар', 'Кувајт',
            'Бахреин', 'Оман', 'Јемен', 'Ирак', 'Иран', 'Авганистан', 'Пакистан',
            'Индија', 'Бангладеш', 'Непал', 'Шри Ланка', 'Малдиви',
            'Кина', 'Јапонија', 'Јужна Кореја', 'Северна Кореја', 'Монголија',
            'Виетнам', 'Тајланд', 'Индонезија', 'Малезија', 'Сингапур',
            'Филипини', 'Мјанмар', 'Камбоџа', 'Лаос', 'Бруней',
            'Казахстан', 'Узбекистан', 'Туркменистан', 'Таџикистан', 'Киргистан',
            'Ерменија', 'Азербејџан', 'Грузија',
            # Африка
            'Египет', 'Либија', 'Тунис', 'Алжир', 'Мароко', 'Мавританија',
            'Сенегал', 'Мали', 'Нигер', 'Нигерија', 'Гана', 'Камерун',
            'Кенија', 'Етиопија', 'Сомалија', 'Уганда', 'Танзанија',
            'Мозамбик', 'Зимбабве', 'Замбија', 'Ангола', 'Конго',
            'Судан', 'Јужен Судан', 'Еритреја', 'Џибути',
            'Мадагаскар', 'Маурициус', 'Сејшели',
            'Јужноафриканска Република', 'Намибија', 'Ботсвана', 'Лесото',
            # Америка
            'Соединети Американски Држави', 'Канада', 'Мексико',
            'Бразил', 'Аргентина', 'Чиле', 'Перу', 'Колумбија', 'Венецуела',
            'Еквадор', 'Боливија', 'Парагвај', 'Уругвај', 'Гвајана',
            'Куба', 'Јамајка', 'Хаити', 'Доминиканска Република',
            'Гватемала', 'Хондурас', 'Ел Салвадор', 'Никарагва', 'Костарика', 'Панама',
            # Океанија
            'Австралија', 'Нов Зеланд', 'Папуа Нова Гвинеја', 'Фиџи',
        ]
        seen = set()
        for c in countries:
            if c not in seen:
                conn.execute('INSERT INTO countries (name) VALUES (?)', (c,))
                seen.add(c)

    for key, value in [
        ('school_name', ''),
        ('municipality', ''),
        ('verification_act', ''),
        ('verification_date', ''),
        ('verification_act_numbers', ''),
        ('verification_issued_by', ''),
        ('ministry_name', ''),
        ('verification_act_entries_json', '[]'),
        ('ministry_entries_json', '[]')
    ]:
        conn.execute('INSERT OR IGNORE INTO settings (key, value) VALUES (?,?)', (key, value))

    if not conn.execute('SELECT id FROM mk_cities').fetchone():
        cities = [
            # Скопје и околина
            'Скопје', 'Арачиново', 'Батинци', 'Бардовци', 'Блаце', 'Бутел', 'Визбегово',
            'Говрлево', 'Горно Количани', 'Горно Лисиче', 'Горно Нерези', 'Горно Свиларе',
            'Грчец', 'Добри Дол', 'Долно Количани', 'Долно Лисиче', 'Долно Нерези',
            'Долно Свиларе', 'Драчево', 'Зелениково', 'Злокуќани', 'Илинден', 'Инџиково',
            'Кучевиште', 'Кучково', 'Лепток', 'Љубанци', 'Љубин', 'Мала Речица', 'Матка',
            'Мирковци', 'Нерези', 'Петровец', 'Побожје', 'Радишани', 'Сарај', 'Сингелиќ',
            'Смилковци', 'Сопиште', 'Студеничани', 'Чучер-Сандево', 'Шуто Оризари',
            'Горно Оризари', 'Долно Оризари', 'Кисела Вода', 'Карпош', 'Аеродром',
            'Центар', 'Гази Баба', 'Ѓорче Петров', 'Чаир', 'Бутел', 'Пагаруша',
            # Битола и Пелагонија
            'Битола', 'Бистрица', 'Братино', 'Брусник', 'Добрушево', 'Долно Орово',
            'Горно Орово', 'Гопеш', 'Граешница', 'Доленци', 'Казани', 'Карамани',
            'Крклино', 'Кукуречани', 'Лавци', 'Лера', 'Лопатица', 'Магарево', 'Метимир',
            'Могила', 'Новаци', 'Облаково', 'Оптичари', 'Породин', 'Ристовац', 'Секирани',
            'Смилево', 'Суводол', 'Трапчин Дол', 'Трн', 'Цапари', 'Цер', 'Чепигово',
            'Тројаци', 'Лознани', 'Канино', 'Довезенце', 'Ладање', 'Дихово',
            # Прилеп
            'Прилеп', 'Алинци', 'Беровци', 'Витолиште', 'Врпско', 'Дебреште', 'Долнени',
            'Зрзе', 'Канатларци', 'Кривогаштани', 'Крушеани', 'Куково', 'Лажани',
            'Мажучиште', 'Небрегово', 'Ореховец', 'Патец', 'Подмол', 'Расоватец',
            'Ропотово', 'Селце', 'Стругово', 'Тополчани', 'Штавица', 'Мамутчево',
            'Бонче', 'Варош', 'Бучин', 'Кален', 'Плетвар', 'Češinovo',
            # Охрид
            'Охрид', 'Вевчани', 'Горно Лакочереј', 'Горно Средорек', 'Горно Стрезево',
            'Долно Стрезево', 'Коњско', 'Косел', 'Лескоец', 'Оровник', 'Пештани',
            'Рамне', 'Сирула', 'Сливово', 'Трпејца', 'Велестово', 'Рашино', 'Куратица',
            'Лактиње', 'Горно Лакочереј', 'Долно Лакочереј', 'Пештани', 'Коњско',
            'Љубаништа', 'Свети Стефан', 'Рамне', 'Климештани', 'Горица',
            # Струга
            'Струга', 'Бичаница', 'Делогожди', 'Дрслајца', 'Лабуниште', 'Октиси',
            'Радолишта', 'Ташмаруниште', 'Франгово', 'Требениште', 'Враниште',
            'Калиште', 'Дренок', 'Локум', 'Мислодежда', 'Подгорци',
            # Дебар и Центар Жупа
            'Дебар', 'Банско', 'Беловиште', 'Горно Косоврасти', 'Долно Косоврасти',
            'Јабланица', 'Кичиница', 'Луково', 'Треbishte', 'Центар Жупа', 'Корошишта',
            'Отиштино', 'Нови Дебар', 'Жировница', 'Требиште', 'Оровник',
            # Кичево
            'Кичево', 'Арбиново', 'Другово', 'Зајас', 'Осломеј', 'Поповјани',
            'Србиново', 'Вранештица', 'Белица', 'Бериково', 'Стрежево', 'Церово',
            'Долна Саlatino', 'Горна Саlатино', 'Речани', 'Велмеј',
            # Гостивар
            'Гостивар', 'Горна Бањица', 'Горно Јеловце', 'Долна Бањица', 'Долно Јеловце',
            'Здуње', 'Корито', 'Мали Влај', 'Ново Село', 'Речане', 'Шемшево',
            'Врапчиште', 'Бозовце', 'Горно Палчиште', 'Долно Палчиште', 'Камењане',
            'Стримница', 'Требош', 'Вруток', 'Челопек', 'Царевик',
            # Тетово
            'Тетово', 'Боговиње', 'Брвеница', 'Желино', 'Јажинце', 'Лешок', 'Теарце',
            'Јажинце', 'Пирок', 'Рамно', 'Брест', 'Мало Речане', 'Голема Речица',
            'Мала Речица', 'Попова Шапка', 'Селце', 'Горно Палчиште', 'Долно Јеловце',
            'Неготино-Полог', 'Фалиш', 'Зелино', 'Прелубиње', 'Долно Коњаре',
            'Горно Коњаре', 'Шипковица', 'Клечовце', 'Одри',
            # Велес
            'Велес', 'Богомила', 'Бузалково', 'Горно Врановци', 'Долно Врановци',
            'Иванковци', 'Чашка', 'Клисура', 'Теово', 'Отиштино', 'Рашани',
            'Горно Јаболково', 'Долно Јаболково', 'Кременица', 'Попадија', 'Отиштино',
            # Куманово
            'Куманово', 'Арбасанци', 'Биљаник', 'Гулинце', 'Клечовце', 'Кокошиње',
            'Конопница', 'Лопате', 'Матеич', 'Младо Нагоричане', 'Никуштак',
            'Орах', 'Романовце', 'Старо Нагоричане', 'Страцин', 'Табановце',
            'Батинце', 'Карбинци', 'Пчиња', 'Ранковце', 'Липково', 'Слупчане', 'Отља',
            'Студена Бара', 'Брест', 'Четирце',
            # Штип
            'Штип', 'Брњарци', 'Горни Балван', 'Долни Балван', 'Карбинци',
            'Лесковица', 'Радање', 'Ново Село', 'Свети Николе', 'Горобинци',
            'Кнежје', 'Малино', 'Мустафино', 'Ораовец', 'Цветово',
            # Кочани и Виница
            'Кочани', 'Виница', 'Горно Градче', 'Долно Градче', 'Зрновци',
            'Мородвис', 'Оризари', 'Пресека', 'Чешиново', 'Облешево',
            'Грљани', 'Лески', 'Трсино', 'Горобинци', 'Еловец',
            # Струмица
            'Струмица', 'Банско', 'Босилово', 'Василево', 'Дабиља', 'Дрен',
            'Колешино', 'Мокриево', 'Ново Село', 'Пиперево', 'Просениково',
            'Сачево', 'Свидовица', 'Сушево', 'Турново', 'Требарово', 'Куклиш',
            'Попчево', 'Муртино', 'Водоча', 'Вељуса', 'Берово',
            # Гевгелија
            'Гевгелија', 'Богородица', 'Конско', 'Лубница', 'Мрзоец', 'Негорци',
            'Прдејци', 'Реца', 'Серменин', 'Стојаково', 'Удово', 'Богданци',
            'Стојаково', 'Гавато', 'Дојран', 'Ново Дојран', 'Старо Дојран',
            # Кавадарци и Неготино
            'Кавадарци', 'Возарци', 'Гарниково', 'Дреново', 'Марена', 'Неготино',
            'Ресава', 'Росоман', 'Сирково', 'Тимјаник', 'Коњух', 'Рајец',
            'Демир Капија', 'Удово', 'Конче', 'Валандово',
            # Берово и Делчево
            'Берово', 'Делчево', 'Будинарци', 'Владимирово', 'Мачево',
            'Митрашинци', 'Русиново', 'Смојмирово', 'Ратево', 'Пехчево',
            'Иловица', 'Робово', 'Саса', 'Чифлик', 'Македонска Каменица',
            # Кратово и Крива Паланка
            'Кратово', 'Крива Паланка', 'Злетово', 'Пробиштип', 'Страцин',
            'Жегљане', 'Псача', 'Мождивњак', 'Петралица', 'Луково',
            # Македонски Брод и Крушево
            'Македонски Брод', 'Крушево', 'Демир Хисар', 'Брезово', 'Журче',
            'Жван', 'Скочивир', 'Поречје', 'Самоков',
            # Ресен
            'Ресен', 'Асамати', 'Долно Перово', 'Евла', 'Завој', 'Козјак',
            'Лева Река', 'Претор', 'Царево Село', 'Евла', 'Бач',
            # Радовиш
            'Радовиш', 'Конче', 'Подареш', 'Ораовица',
            # Маврово и Ростуше
            'Маврово', 'Нивиште', 'Ростуше', 'Леуново', 'Лазарополе', 'Галичник',
            'Гари', 'Бистра', 'Тресонче', 'Маврови Анови', 'Жировница', 'Никифорово',
            'Сретково', 'Волковија',
            # Лозово / Чашка / Градско
            'Лозово', 'Градско', 'Богомила', 'Рамно',
            # Зрновци
            'Зрновци', 'Горобинци', 'Еловец',
            # Свети Николе
            'Свети Николе', 'Кнежје', 'Малино', 'Мустафино', 'Ораовец', 'Цветово',
            # Разни
            'Јегуновце', 'Теарце', 'Боговиње', 'Желино', 'Зелино',
            'Старо Нагоричане', 'Млaдо Нагоричане', 'Ранковце',
            'Чашка', 'Кривогаштани', 'Долнени', 'Могила',
            'Новаци', 'Прилеп', 'Демир Хисар', 'Ресен',
        ]
        seen = set()
        for city in cities:
            if city not in seen:
                conn.execute('INSERT OR IGNORE INTO mk_cities (name) VALUES (?)', (city,))
                seen.add(city)

    conn.commit()
    conn.close()


# --- Auth ---

def get_user_by_id(user_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()
    return User(row['id'], row['username'], row['role'], row['full_name']) if row else None


def authenticate_user(username, password):
    conn = get_db()
    row = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
    conn.close()
    if row and check_password_hash(row['password_hash'], password):
        return User(row['id'], row['username'], row['role'], row['full_name'])
    return None


# --- Settings ---

def get_settings():
    conn = get_db()
    rows = conn.execute('SELECT key, value FROM settings').fetchall()
    conn.close()
    return {r['key']: r['value'] for r in rows}


def save_settings(form):
    conn = get_db()
    for key, value in form.items():
        conn.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)', (key, value))
    conn.commit()
    conn.close()


# --- Classes ---

def get_all_classes():
    conn = get_db()
    rows = conn.execute('SELECT * FROM classes ORDER BY name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_user_classes(user_id, role):
    conn = get_db()
    if role == 'admin':
        rows = conn.execute('SELECT * FROM classes ORDER BY name').fetchall()
    else:
        rows = conn.execute('''
            SELECT c.* FROM classes c
            JOIN user_classes uc ON c.id = uc.class_id
            WHERE uc.user_id = ? ORDER BY c.name
        ''', (user_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_class(name):
    conn = get_db()
    conn.execute('INSERT OR IGNORE INTO classes (name) VALUES (?)', (name,))
    conn.commit()
    conn.close()


def delete_class(class_id):
    conn = get_db()
    conn.execute('DELETE FROM classes WHERE id = ?', (class_id,))
    conn.commit()
    conn.close()


# --- Students ---

def get_all_students():
    conn = get_db()
    rows = conn.execute('''
        SELECT s.*, c.name as class_name FROM students s
        LEFT JOIN classes c ON s.class_id = c.id
        ORDER BY c.name, s.last_name, s.first_name
    ''').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_students_by_class(class_id):
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM students WHERE class_id = ? ORDER BY last_name, first_name',
        (class_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_student(student_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM students WHERE id = ?', (student_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def add_student(form):
    conn = get_db()
    conn.execute('''
        INSERT INTO students
        (first_name, last_name, parent_name, birth_date, birth_place,
         birth_municipality, birth_country, citizenship, class_id)
        VALUES (?,?,?,?,?,?,?,?,?)
    ''', (
        form.get('first_name'), form.get('last_name'), form.get('parent_name'),
        form.get('birth_date'), form.get('birth_place'), form.get('birth_municipality'),
        form.get('birth_country', 'Република Северна Македонија'),
        form.get('citizenship', 'македонско'), form.get('class_id') or None
    ))
    conn.commit()
    conn.close()


def update_student(student_id, form):
    conn = get_db()
    conn.execute('''
        UPDATE students SET first_name=?, last_name=?, parent_name=?, birth_date=?,
        birth_place=?, birth_municipality=?, birth_country=?, citizenship=?, class_id=?
        WHERE id=?
    ''', (
        form.get('first_name'), form.get('last_name'), form.get('parent_name'),
        form.get('birth_date'), form.get('birth_place'), form.get('birth_municipality'),
        form.get('birth_country'), form.get('citizenship'), form.get('class_id') or None,
        student_id
    ))
    conn.commit()
    conn.close()


def delete_student(student_id):
    conn = get_db()
    conn.execute('DELETE FROM certificates WHERE student_id = ?', (student_id,))
    conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
    conn.commit()
    conn.close()


def clear_students():
    conn = get_db()
    conn.execute('DELETE FROM certificates WHERE student_id IS NOT NULL')
    conn.execute('DELETE FROM students')
    conn.commit()
    conn.close()


def get_duplicate_students():
    students = get_all_students()
    groups = {}
    for s in students:
        key = (s.get('first_name'), s.get('last_name'), s.get('parent_name'),
               s.get('birth_date'), s.get('birth_place'), s.get('birth_municipality'),
               s.get('birth_country'), s.get('citizenship'), s.get('class_id'))
        groups.setdefault(key, []).append(s)
    return [sorted(g, key=lambda s: s['id']) for g in groups.values() if len(g) > 1]


def delete_students(student_ids):
    if not student_ids:
        return
    conn = get_db()
    conn.executemany('DELETE FROM certificates WHERE student_id = ?', [(sid,) for sid in student_ids])
    conn.executemany('DELETE FROM students WHERE id = ?', [(sid,) for sid in student_ids])
    conn.commit()
    conn.close()


def export_students_xlsx():
    students = get_all_students()
    wb = Workbook()
    ws = wb.active
    ws.append(STUDENT_XLSX_HEADERS)
    for s in students:
        ws.append([
            s.get('last_name'), s.get('first_name'), s.get('parent_name'),
            s.get('class_name'), s.get('birth_date'), s.get('birth_place'),
            s.get('birth_municipality'), s.get('birth_country'), s.get('citizenship')
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_students_xlsx(file_stream):
    wb = load_workbook(filename=file_stream, data_only=True)
    ws = wb.active
    classes = {_normalize_class_name(c['name']): c['id'] for c in get_all_classes()}
    header_names = {h.strip().lower() for h in STUDENT_XLSX_HEADERS}

    imported = 0
    skipped = 0
    unmatched = set()

    conn = get_db()
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        row = list(row) + [None] * (9 - len(row))
        (last_name, first_name, parent_name, class_name, birth_date,
         birth_place, birth_municipality, birth_country, citizenship) = row[:9]

        if not first_name or not last_name:
            skipped += 1
            continue

        if str(first_name).strip().lower() in header_names and str(last_name).strip().lower() in header_names:
            skipped += 1
            continue

        class_id = None
        if class_name:
            class_name = _normalize_class_name(class_name)
            class_id = classes.get(class_name)
            if class_id is None:
                unmatched.add(class_name)

        conn.execute('''
            INSERT INTO students
            (first_name, last_name, parent_name, birth_date, birth_place,
             birth_municipality, birth_country, citizenship, class_id)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (
            str(first_name).strip(), str(last_name).strip(),
            str(parent_name).strip() if parent_name else None,
            str(birth_date).strip() if birth_date else None,
            str(birth_place).strip() if birth_place else None,
            str(birth_municipality).strip() if birth_municipality else None,
            str(birth_country).strip() if birth_country else 'Република Северна Македонија',
            str(citizenship).strip() if citizenship else 'македонско',
            class_id
        ))
        imported += 1

    conn.commit()
    conn.close()
    return imported, skipped, sorted(unmatched), None, None


# --- Subjects ---

def get_all_subjects():
    conn = get_db()
    rows = conn.execute('SELECT * FROM subjects ORDER BY struka, profil, name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_subject(form):
    conn = get_db()
    conn.execute(
        'INSERT INTO subjects (name, struka, profil) VALUES (?,?,?)',
        (form.get('name'), form.get('struka') or None, form.get('profil') or None)
    )
    conn.commit()
    conn.close()


def delete_subject(subject_id):
    conn = get_db()
    conn.execute('DELETE FROM subjects WHERE id = ?', (subject_id,))
    conn.commit()
    conn.close()


# --- Ministries ---

def get_all_ministries():
    conn = get_db()
    rows = conn.execute('SELECT * FROM ministries ORDER BY name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_ministry(form):
    conn = get_db()
    conn.execute(
        'INSERT INTO ministries (name, act_number, act_date) VALUES (?,?,?)',
        (form.get('name'), form.get('act_number') or None, form.get('act_date') or None)
    )
    conn.commit()
    conn.close()


def delete_ministry(ministry_id):
    conn = get_db()
    conn.execute('DELETE FROM ministries WHERE id = ?', (ministry_id,))
    conn.commit()
    conn.close()


# --- Behavior / Countries ---

def get_behavior_types():
    conn = get_db()
    rows = conn.execute('SELECT * FROM behavior_types ORDER BY id').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_countries():
    conn = get_db()
    rows = conn.execute('SELECT * FROM countries ORDER BY name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_mk_cities():
    conn = get_db()
    rows = conn.execute('SELECT name FROM mk_cities ORDER BY name').fetchall()
    conn.close()
    return [r['name'] for r in rows]


# --- Users ---

def get_all_users():
    conn = get_db()
    users = conn.execute(
        'SELECT id, username, role, full_name FROM users ORDER BY role, full_name'
    ).fetchall()
    result = []
    for u in users:
        u = dict(u)
        classes = conn.execute('''
            SELECT c.id, c.name FROM classes c
            JOIN user_classes uc ON c.id = uc.class_id
            WHERE uc.user_id = ?
        ''', (u['id'],)).fetchall()
        u['classes'] = [dict(c) for c in classes]
        result.append(u)
    conn.close()
    return result


def add_user(form):
    conn = get_db()
    conn.execute(
        'INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)',
        (form.get('username'), generate_password_hash(form.get('password')),
         form.get('role', 'teacher'), form.get('full_name'))
    )
    conn.commit()
    conn.close()


def delete_user(user_id):
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()


def change_password(user_id, new_password):
    conn = get_db()
    conn.execute(
        'UPDATE users SET password_hash = ? WHERE id = ?',
        (generate_password_hash(new_password), user_id)
    )
    conn.commit()
    conn.close()


def assign_class_to_user(user_id, class_id):
    conn = get_db()
    conn.execute(
        'INSERT OR IGNORE INTO user_classes (user_id, class_id) VALUES (?,?)',
        (user_id, class_id)
    )
    conn.commit()
    conn.close()


def remove_class_from_user(user_id, class_id):
    conn = get_db()
    conn.execute(
        'DELETE FROM user_classes WHERE user_id = ? AND class_id = ?',
        (user_id, class_id)
    )
    conn.commit()
    conn.close()


# --- Certificates ---

def _parse_subjects(form):
    subjects = []
    i = 0
    while True:
        name = form.get(f'subject_{i}')
        if name is None:
            break
        if name.strip():
            subjects.append({
                'name': name.strip(),
                'grade_text': form.get(f'grade_text_{i}', ''),
                'grade_num': form.get(f'grade_num_{i}', ''),
            })
        i += 1
    return subjects


def save_certificate(form, user_id):
    subjects = _parse_subjects(form)
    conn = get_db()
    cur = conn.execute('''
        INSERT INTO certificates (
            student_id, student_name, class_id, school_year, attempt_number, completed_year_top, class_year,
            parent_name, birth_date, birth_place, birth_municipality, birth_country, citizenship,
            main_book_number, main_book_year, subjects_grades, behavior, absences_justified, absences_unjustified,
            completed_year, education_type, area, profile, overall_grade,
            place, cert_date, del_br, verification_act, verification_date, verification_issued_by,
            class_teacher, director,
            status, created_by
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        form.get('student_id') or None, form.get('student_name'),
        form.get('class_id') or None, form.get('school_year'),
        form.get('attempt_number'), form.get('completed_year_top'), form.get('class_year'),
        form.get('parent_name'), form.get('birth_date'), form.get('birth_place'),
        form.get('birth_municipality'), form.get('birth_country'), form.get('citizenship'),
        form.get('main_book_number'), form.get('main_book_year'), json.dumps(subjects, ensure_ascii=False),
        form.get('behavior'), form.get('absences_justified'), form.get('absences_unjustified'),
        form.get('completed_year'), form.get('education_type'), form.get('area'),
        form.get('profile'), form.get('overall_grade'),
        form.get('place'), form.get('cert_date'), form.get('del_br'),
        form.get('verification_act'), form.get('verification_date'), form.get('verification_issued_by'),
        form.get('class_teacher'), form.get('director'),
        form.get('status', 'draft'), user_id
    ))
    cert_id = cur.lastrowid
    conn.commit()
    conn.close()
    return cert_id


def update_certificate(cert_id, form):
    subjects = _parse_subjects(form)
    conn = get_db()
    conn.execute('''
        UPDATE certificates SET
            student_id=?, student_name=?, class_id=?, school_year=?, attempt_number=?, completed_year_top=?, class_year=?,
            parent_name=?, birth_date=?, birth_place=?, birth_municipality=?, birth_country=?, citizenship=?,
            main_book_number=?, main_book_year=?, subjects_grades=?, behavior=?, absences_justified=?, absences_unjustified=?,
            completed_year=?, education_type=?, area=?, profile=?, overall_grade=?,
            place=?, cert_date=?, del_br=?, verification_act=?, verification_date=?, verification_issued_by=?,
            class_teacher=?, director=?,
            status=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    ''', (
        form.get('student_id') or None, form.get('student_name'),
        form.get('class_id') or None, form.get('school_year'),
        form.get('attempt_number'), form.get('completed_year_top'), form.get('class_year'),
        form.get('parent_name'), form.get('birth_date'), form.get('birth_place'),
        form.get('birth_municipality'), form.get('birth_country'), form.get('citizenship'),
        form.get('main_book_number'), form.get('main_book_year'), json.dumps(subjects, ensure_ascii=False),
        form.get('behavior'), form.get('absences_justified'), form.get('absences_unjustified'),
        form.get('completed_year'), form.get('education_type'), form.get('area'),
        form.get('profile'), form.get('overall_grade'),
        form.get('place'), form.get('cert_date'), form.get('del_br'),
        form.get('verification_act'), form.get('verification_date'), form.get('verification_issued_by'),
        form.get('class_teacher'), form.get('director'),
        form.get('status', 'draft'), cert_id
    ))
    conn.commit()
    conn.close()


def get_certificate(cert_id):
    conn = get_db()
    row = conn.execute('''
        SELECT c.*, cl.name as class_name FROM certificates c
        LEFT JOIN classes cl ON c.class_id = cl.id
        WHERE c.id = ?
    ''', (cert_id,)).fetchone()
    conn.close()
    if not row:
        return None
    cert = dict(row)
    cert['subjects_grades'] = json.loads(cert['subjects_grades'] or '[]')
    if not cert.get('completed_year_top'):
        cert['completed_year_top'] = cert.get('completed_year') or ''

    # Fallback for legacy rows created before these fields were stored in certificates.
    if cert.get('student_id') and any(not cert.get(k) for k in [
        'parent_name', 'birth_date', 'birth_place', 'birth_municipality', 'birth_country', 'citizenship'
    ]):
        conn = get_db()
        student = conn.execute('''
            SELECT parent_name, birth_date, birth_place, birth_municipality, birth_country, citizenship
            FROM students WHERE id = ?
        ''', (cert['student_id'],)).fetchone()
        conn.close()
        if student:
            student = dict(student)
            for key, value in student.items():
                if not cert.get(key):
                    cert[key] = value

    return cert


def get_certificates(user_id, role, class_filter=None, status_filter='all'):
    conn = get_db()
    query = '''
        SELECT c.*, cl.name as class_name FROM certificates c
        LEFT JOIN classes cl ON c.class_id = cl.id WHERE 1=1
    '''
    params = []
    if role != 'admin':
        query += ' AND c.class_id IN (SELECT class_id FROM user_classes WHERE user_id = ?)'
        params.append(user_id)
    if class_filter:
        query += ' AND c.class_id = ?'
        params.append(class_filter)
    if status_filter != 'all':
        query += ' AND c.status = ?'
        params.append(status_filter)
    query += ' ORDER BY c.updated_at DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    result = []
    for row in rows:
        r = dict(row)
        r['subjects_grades'] = json.loads(r['subjects_grades'] or '[]')
        result.append(r)
    return result


def delete_certificate(cert_id):
    conn = get_db()
    conn.execute('DELETE FROM certificates WHERE id = ?', (cert_id,))
    conn.commit()
    conn.close()


def get_stats():
    conn = get_db()
    stats = {
        'students': conn.execute('SELECT COUNT(*) FROM students').fetchone()[0],
        'certificates': conn.execute('SELECT COUNT(*) FROM certificates').fetchone()[0],
        'ready': conn.execute("SELECT COUNT(*) FROM certificates WHERE status='ready'").fetchone()[0],
        'draft': conn.execute("SELECT COUNT(*) FROM certificates WHERE status='draft'").fetchone()[0],
        'users': conn.execute('SELECT COUNT(*) FROM users').fetchone()[0],
        'classes': conn.execute('SELECT COUNT(*) FROM classes').fetchone()[0],
    }
    conn.close()
    return stats
